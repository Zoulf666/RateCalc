import model
import time

from openpyxl import load_workbook
from math import ceil


def calc_price(weight, first_weight_num, first_weight_price, next_weight_price):
    weight = ceil(weight)  # 向上取整
    if weight <= first_weight_num:
        price = first_weight_price
    else:
        remain = ceil(weight - first_weight_num)
        remain_price = remain * next_weight_price
        price = first_weight_price + remain_price
    return float(price)


def excel_handle(path):
    start = time.time()
    # 注：openpyxl 下标从1开始
    wb = load_workbook(path)
    sheets = wb.worksheets
    custom = model.Custom()
    custom_detail = model.CustomDetail()
    unsign_info = {}

    for sheet in sheets:
        price_col = sheet.max_column - 1
        price_sum_col = sheet.max_column
        max_row = sheet.max_row
        sum_price = 0

        # 去除空sheet
        if max_row <= 1:
            continue

        # 在最后一行增加运费与运费总数
        if not sheet.cell(row=1, column=price_sum_col).value == '运费总数':
            price_col = sheet.max_column + 1
            price_sum_col = price_col + 1
            sheet.cell(row=1, column=price_col).value = '运费'
            sheet.cell(row=1, column=price_sum_col).value = '运费总数'

        # 动态表头位置
        head = {}
        for col in range(1, sheet.max_column):
            if sheet.cell(1, col).value == '重量':
                head['weight'] = col
            if sheet.cell(1, col).value == '目的省份':
                head['provice'] = col
            if sheet.cell(1, col).value == '寄件客户':
                head['name'] = col

        # 检查表头格式
        for _, v in head.items():
            if not v:
                raise Exception('sheet - {} 表头名称错误，请检查表头是否为 重量/ 目的省份/ 寄件客户!'.format(sheet.title))

        # 姓名集合（集合特性去重）
        custom_names = set(v.value for index, v in enumerate(sheet[chr(64 + head['name'])]) if not index == 0)
        print(custom_names)

        center = time.time()
        print('Center1: {}'.format(center - start))

        """
        name_dict字典格式：{
                                    '张三': {
                                                '湖南': {
                                                                '首重重量': 1.0,
                                                                ...                                            
                                                           }
                                                '湖北': ...
                                                }
                                    '李四': ...
                             }
        """
        name_dict = {}
        for custom_name in custom_names:
            custom_id = custom.find_custom_like(custom_name)
            # 没找到指定用户（custom_id=0）抛出异常
            if not custom_id:
                # raise Exception('sheet - {}, 用户名: {} 未录入数据库中！'.format(sheetname, custom_name))
                unsign_info[custom_name] = 'sheet-{}'.format(sheet.title)
            else:
                name_dict[custom_name] = custom_detail.fetch_custom_detail(custom_id)
        print(name_dict)

        center = time.time()
        print('Center2: {}'.format(center - start))

        # 从第2行开始为数据
        for row in range(2, max_row + 1):
            weight = sheet.cell(row=row, column=head['weight']).value
            provice = sheet.cell(row=row, column=head['provice']).value
            name = sheet.cell(row=row, column=head['name']).value

            # 检查名字是否存在于数据库汇总
            if name not in name_dict:
                continue

            # 模糊匹配省份
            can_calc = False
            if provice in name_dict[name]:
                can_calc = True
            else:
                custom_id = custom.find_custom_like(name)
                if custom_detail.find_custom_detail(custom_id, provice):
                    provice = custom_detail.find_provice(custom_id, provice)
                    if provice:
                        can_calc = True

            if can_calc:
                first_weight_num = name_dict[name][provice]['首重重量']
                first_weight_price = name_dict[name][provice]['首重价格']
                next_weight_price = name_dict[name][provice]['续重价格']
                price = calc_price(weight, first_weight_num, first_weight_price, next_weight_price)
                sum_price += price
                sheet.cell(row=row, column=price_col).value = price
            else:
                raise Exception('sheet - {}, 第 {} 行, 目的省份{}未录入数据库中！'.format(sheet.title, row, provice))
        # 写入总金额
        sheet.cell(2, price_sum_col).value = float(sum_price)
    wb.save(path)
    custom.close()
    custom_detail.close()
    end = time.time()
    print(end - start)
    return unsign_info


def excel_provice_handle(path):
    wb = load_workbook(path)
    sheetnames = wb.sheetnames

    custom = model.Custom()
    custom_detail = model.CustomDetail()
    for name in sheetnames:
        # 判断客户是否存在，不存在创建
        custom_id = custom.find_custom(name)
        if custom_id == 0:
            custom_id = custom.add_custom(name)
        sheet = wb[name]

        header1 = sheet.cell(row=1, column=1).value
        header2 = sheet.cell(row=1, column=2).value
        header3 = sheet.cell(row=1, column=3).value
        header4 = sheet.cell(row=1, column=4).value
        if not header1 == '目的地' and not header2 == '首重重量' and not header3 == '首重价格' and not header4 == '续重价格':
            custom.delete_custom(custom_id)
            raise Exception('表头格式错误！请按照样例格式输入！')

        for row in range(2, sheet.max_row + 1):
            provice = sheet.cell(row=row, column=1).value
            # 因为可能把空的单元格算进去，所以需要判断
            if not provice:
                continue
            f_num = float(sheet.cell(row=row, column=2).value)
            f_price = float(sheet.cell(row=row, column=3).value)
            n_price = float(sheet.cell(row=row, column=4).value)

            if custom_detail.find_custom_detail(custom_id, provice):
                custom_detail.update_custom_detail(custom_id, provice, f_num, f_price, n_price)
            else:
                custom_detail.add_custom_detail(custom_id, provice, f_num, f_price, n_price)
    custom.close()
    custom_detail.close()


def excel_handle2(path):
    # 注：openpyxl 下标从1开始
    wb = load_workbook(path)
    sheets = wb.worksheets
    custom = model.Custom()
    custom_detail = model.CustomDetail()
    unsign_info = {}  # 记录未注册用户 name: sheet位置
    head = {}  # 表头位置
    for sheet in sheets:
        price_col = sheet.max_column - 1
        price_sum_col = sheet.max_column
        max_row = sheet.max_row
        sum_price = 0

        # 去除空sheet
        if max_row <= 1:
            continue

        # 在最后一行增加运费与运费总数
        if not sheet.cell(row=1, column=price_sum_col).value == '运费总数':
            price_col = sheet.max_column + 1
            price_sum_col = price_col + 1
            sheet.cell(row=1, column=price_col).value = '运费'
            sheet.cell(row=1, column=price_sum_col).value = '运费总数'

        # 动态获取表头位置并检查
        for col in range(1, sheet.max_column):
            val = sheet.cell(1, col).value
            if val == '重量':
                head['weight'] = col
            elif val == '目的省份':
                head['provice'] = col
            elif val == '寄件客户':
                head['name'] = col
            else:
                raise Exception('sheet - {} 表头名称错误，请检查表头是否为 重量/ 目的省份/ 寄件客户!'.format(sheet.title))

        # 获取所有客户
        custom_names = [name for name, _ in enumerate(custom.fetch_custom())]
        """
        name_dict字典格式：{
                                    '张三': {
                                                '湖南': {
                                                                '首重重量': 1.0,
                                                                ...                                            
                                                           }
                                                '湖北': ...
                                                }
                                    '李四': ...
                             }
        """
        name_dict = {}
        for custom_name in custom_names:
            custom_id = custom.find_custom(custom_name)
            name_dict[custom_name] = custom_detail.fetch_custom_detail(custom_id)

        # 从第2行开始为数据
        for row in range(2, max_row + 1):
            weight = sheet.cell(row=row, column=head['weight']).value
            provice = sheet.cell(row=row, column=head['provice']).value
            name = sheet.cell(row=row, column=head['name']).value

            # 检查名字是否存在于数据库汇总
            if name not in custom_names:
                unsign_info[name] = 'Sheet-{}'.format(sheet.title)
                continue

            # 模糊匹配省份
            can_calc = False
            if provice in name_dict[name]:
                can_calc = True
            else:
                custom_id = custom.find_custom_like(name)
                if custom_detail.find_custom_detail(custom_id, provice):
                    provice = custom_detail.find_provice(custom_id, provice)
                    if provice:
                        can_calc = True

            if can_calc:
                first_weight_num = name_dict[name][provice]['首重重量']
                first_weight_price = name_dict[name][provice]['首重价格']
                next_weight_price = name_dict[name][provice]['续重价格']
                price = calc_price(weight, first_weight_num, first_weight_price, next_weight_price)
                sum_price += price
                sheet.cell(row=row, column=price_col).value = price
            else:
                raise Exception('sheet - {}, 第 {} 行, 目的省份{}未录入数据库中！'.format(sheet.title, row, provice))
        # 写入总金额
        sheet.cell(2, price_sum_col).value = float(sum_price)

    wb.save(path)
    custom.close()
    custom_detail.close()
    return unsign_info


if __name__ == '__main__':
    start = time.time()
    # usign = excel_handle2('/home/zoulf/mycode/RateCalc/9月做账定.xlsx')
    wb = load_workbook('/home/zoulf/mycode/RateCalc/9月做账定.xlsx')
    end = time.time()
    print(end - start)
    # print(usign)