import model

from xlrd import open_workbook
from openpyxl import load_workbook
from json import dumps
from json import loads
from math import ceil


def calc_special_custom_price(weight, price_dict):
    """
    :param weight: float
    :param price_dict: {'<=3':2.6, '<=2': 1.6 ...}
    :return: price
    """
    price = 0
    extra_price = 0
    lower_dict = {}
    for key, value in price_dict.items():
        if '<=' in key:
            num = float(key.split('<=')[1])
            lower_dict[num] = float(value)
        if '>' in key:
            extra_price = float(value)
    lower_list = sorted(lower_dict.keys(), reverse=True)
    if weight > lower_list[0]:
        price = lower_dict[lower_list[0]] + (weight - lower_list[0]) * extra_price
    else:
        for k, num in enumerate(lower_list):
            if weight > num:
                price = lower_dict[lower_list[k - 1]]
                break
            else:
                price = lower_dict[lower_list[-1]]

    return price


def calc_price(weight, first_weight_num, first_weight_price, next_weight_price):
    weight = ceil(weight)  # 向上取整
    if weight <= first_weight_num:
        price = first_weight_price
    else:
        remain = ceil(weight - first_weight_num)
        remain_price = remain * next_weight_price
        price = first_weight_price + remain_price
    return float(price)


def excel_provice_handle(path):
    data = open_workbook(path)
    sheets = data.sheets()

    custom = model.Custom()
    custom_detail = model.CustomDetail()
    special_custom = model.SpecialCustomDetail()
    new_customs = []
    error_customs = []
    for sheet in sheets:
        sheetname = sheet.name
        first_rows = sheet.row_values(0)
        tr_index_dict = {}
        is_error_date = True
        is_special = 0

        # 遍历第一行（表头）的值
        for index, value in enumerate(first_rows):
            if value:
                value = value.strip()
            else:
                continue

            if value == '目的地':
                tr_index_dict['目的地'] = index
            elif value == '首重重量':
                tr_index_dict['首重重量'] = index
            elif value == '首重价格':
                tr_index_dict['首重价格'] = index
            elif value == '续重价格':
                tr_index_dict['续重价格'] = index
            elif '<=' in value:
                value = '<=' + value.split('<=')[1]
                tr_index_dict[value] = index
            elif '>' in value:
                value = '>' + value.split('>')[1]
                tr_index_dict[value] = index

        # 验证表格结构
        if '目的地' in tr_index_dict:
            if '首重重量' in tr_index_dict and '续重价格' in tr_index_dict:
                is_error_date = False
            else:
                tmp = []
                for k in tr_index_dict.keys():
                    if '<=' in k:
                        is_error_date = False
                    if '>' in k:
                        tmp.append(k)
                # 需要保证同时存在 <= 和 >
                if not tmp:
                    is_error_date = True
                is_special = 1

        if is_error_date:
            error_customs.append(sheetname)
            continue

        custom_id = custom.find_custom(sheetname)
        if custom_id == 0:
            custom_id = custom.add_custom(sheetname, is_special)
            new_customs.append(sheetname)

        # 找到该用户的详情(如果有)，会根据是否是特殊用户来进行判断
        if is_special:
            detail_dict = special_custom.fetch_custom_detail(custom_id)
        else:
            detail_dict = custom_detail.fetch_custom_detail(custom_id)

        # 一行一行的进行遍历
        for row in range(1, sheet.nrows):
            provice = sheet.cell_value(rowx=row, colx=tr_index_dict['目的地'])
            # 普通用户下的操作
            if not is_special:
                # 因为可能把空的单元格算进去，所以需要判断
                if not provice:
                    continue
                try:
                    f_num = float(sheet.cell_value(rowx=row, colx=tr_index_dict['首重重量']))
                    f_price = float(sheet.cell_value(rowx=row, colx=tr_index_dict['首重价格']))
                    n_price = float(sheet.cell_value(rowx=row, colx=tr_index_dict['续重价格']))
                except ValueError:
                    raise ValueError('请检查表 - {} 中的第{}行，确保首重重量/ 首重价格/ 续重价格中不包含中文！'.format(sheetname, row))
                custom_detail_id = custom_detail.find_custom_detail(custom_id, provice)
                # 找到则执行更新操作（小优化的），未找到则执行添加操作
                if custom_detail_id:
                    if provice in detail_dict:
                        if f_num == detail_dict[provice]['首重重量'] \
                                and f_price == detail_dict[provice]['首重价格'] \
                                and n_price == detail_dict[provice]['续重价格']:
                            continue
                        else:
                            custom_detail.update_custom_detail(custom_id, provice, f_num, f_price, n_price)
                    else:
                        custom_detail.update_custom_detail(custom_id, provice, f_num, f_price, n_price)
                else:
                    custom_detail.add_custom_detail(custom_id, provice, f_num, f_price, n_price)
            # 特殊用户下的操作
            else:
                range_dict = {}
                for key in tr_index_dict.keys():
                    if key == '目的地':
                        continue
                    try:
                        range_dict[key] = float(sheet.cell_value(rowx=row, colx=tr_index_dict[key]))
                    except ValueError:
                        raise ValueError('请检查表 - {} 中的第{}行，确保首重重量/ 首重价格/ 续重价格中不包含中文！'.format(sheetname, row))
                range_dict = dumps(range_dict)
                special_custom_id = special_custom.find_custom_detail(custom_id, provice)
                if special_custom_id:
                    if provice in detail_dict:
                        if range_dict == detail_dict[provice]['重量价格']:
                            continue
                        else:
                            special_custom.update_custom_detail(custom_id, provice, range_dict)
                    else:
                        special_custom.update_custom_detail(custom_id, provice, range_dict)
                else:
                    special_custom.add_custom_detail(custom_id, provice, range_dict)

    custom.close()
    custom_detail.close()
    special_custom.close()
    return new_customs, error_customs


def excel_handle2(path):
    # 注：openpyxl 下标从1开始
    wb = load_workbook(path)
    sheetnames = wb.sheetnames
    custom = model.Custom()
    custom_detail = model.CustomDetail()
    special_custom = model.SpecialCustomDetail()
    unsign_info = {}  # 记录未注册用户 name: sheet位置
    error_provice = {}
    head = {}  # 表头位置
    for sheetname in sheetnames:
        current_sheetname = sheetname
        sheet = wb[sheetname]
        price_col = sheet.max_column - 1
        price_sum_col = sheet.max_column
        max_row = sheet.max_row
        sum_price = 0

        # 去除空sheet
        if max_row <= 1:
            continue
        print(sheetname)
        # 动态获取表头位置
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(1, col).value
            if val:
                val = val.strip()
            if val == '重量':
                head['weight'] = col
            elif val == '目的省份':
                head['provice'] = col
            elif val == '寄件客户':
                head['name'] = col
        # 检查表头格式
        if len(head.keys()) != 3:
            raise Exception('sheet 名：{} 表头名称错误，请检查表头是否为 重量/ 目的省份/ 寄件客户!'.format(sheetname))

        # 在最后一行增加运费与运费总数
        if not sheet.cell(row=1, column=price_sum_col).value == '运费总数':
            price_col = sheet.max_column + 1
            price_sum_col = price_col + 1
            sheet.cell(row=1, column=price_col).value = '运费'
            sheet.cell(row=1, column=price_sum_col).value = '运费总数'

        # 获取所有客户
        custom_names = [name for name, _ in custom.fetch_custom().items()]
        """
        name_dict字典格式：{
                                    '张三': {
                                                '湖南': {
                                                                '首重重量': 1.0,
                                                                ...                                            
                                                           }
                                                '湖北': ...
                                                }
                                    '李四': ...
                             }
        """
        name_dict = {}
        for custom_name in custom_names:
            custom_id = custom.find_custom(custom_name)
            is_special = custom.get_special_value(custom_id)
            if is_special:
                name_dict[custom_name] = special_custom.fetch_custom_detail(custom_id)
            else:
                name_dict[custom_name] = custom_detail.fetch_custom_detail(custom_id)

        # 从第2行开始，一行行的数据
        for row in range(2, max_row + 1):
            weight = sheet.cell(row=row, column=head['weight']).value
            provice = sheet.cell(row=row, column=head['provice']).value
            name = sheet.cell(row=row, column=head['name']).value
            is_special = 0
            can_calc = False

            custom_id = custom.find_custom_like(name)
            if custom_id:
                is_special = custom.get_special_value(custom_id)

            # TODO 重量float
            if weight:
                try:
                    weight = float(weight)
                except ValueError:
                    raise ValueError('重量行 - {} 中不能包含中文! '.format(row))
            else:
                continue

            if provice:
                provice = provice.strip()
                if '=' in provice:
                    provice = provice.replace('=', '')
                    provice = sheet[provice].value
            else:
                continue

            if name:
                name = name.strip()
            else:
                continue

            # 检查名字是否存在于数据库中
            if name not in custom_names:
                if name not in unsign_info:
                    unsign_info[name] = ['{}, 第一次出现在工作表<{}> - {}行)'.format(name, current_sheetname, row - 1), 1]
                else:
                    unsign_info[name][-1] += 1
                continue

            # 精确查找未找到时，模糊匹配省份
            if provice in name_dict[name]:
                can_calc = True
            else:
                if is_special:
                    if special_custom.find_custom_detail(custom_id, provice):
                        provice = special_custom.find_provice(custom_id, provice)
                        if provice:
                            can_calc = True
                else:
                    if custom_detail.find_custom_detail(custom_id, provice):
                        provice = custom_detail.find_provice(custom_id, provice)
                        if provice:
                            can_calc = True

            if can_calc:
                if is_special:
                    price_data = name_dict[name][provice]['重量价格']
                    price_dict = loads(price_data)
                    price = calc_special_custom_price(weight, price_dict)
                else:
                    first_weight_num = name_dict[name][provice]['首重重量']
                    first_weight_price = name_dict[name][provice]['首重价格']
                    next_weight_price = name_dict[name][provice]['续重价格']
                    price = calc_price(weight, first_weight_num, first_weight_price, next_weight_price)
                sum_price += price
                sheet.cell(row=row, column=price_col).value = price
            else:
                print(provice)
                error_provice[provice] = row
        # 写入总金额
        sheet.cell(2, price_sum_col).value = float(sum_price)
    wb.save(path)
    custom.close()
    custom_detail.close()
    return unsign_info, error_provice


if __name__ == '__main__':
    # new, error = excel_provice_handle('客户价格表(正确格式数据)2.xlsx')
    # print(new)
    # print(error)
    unsign, error = excel_handle2('11月做账定系统测试.xlsx')
    print(unsign)
    print(error)
    print('====')
